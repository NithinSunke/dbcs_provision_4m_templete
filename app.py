import os
import re
import json
import ntpath
import configparser
import threading
import uuid
from datetime import datetime
from typing import Any

from dotenv import load_dotenv
from flask import Flask, redirect, render_template, request, url_for
from openpyxl import load_workbook

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB upload limit

# In-memory store for demo/small-team use.
# For production, replace with DB persistence (SQLite/PostgreSQL).
BATCHES: dict[str, dict[str, Any]] = {}
BATCH_LOCK = threading.Lock()
DATA_DIR = os.path.join(os.getcwd(), "data")
STORE_PATH = os.path.join(DATA_DIR, "batches.json")

REQUIRED_COLUMNS = [
    "display_name",
    "compartment_id",
    "subnet_id",
    "shape",
    "database_edition",
    "db_version",
    "db_name",
    "admin_password",
    "cpu_core_count",
    "data_storage_size_in_gbs",
]

OCID_PATTERN = re.compile(r"^ocid1\.[a-z0-9._-]+\..+$", re.IGNORECASE)

KV_TO_CANONICAL_MAP = {
    "db_system_name": "display_name",
    "compartment": "compartment_id",
    "compartment_ocid": "compartment_id",
    "availability_domain": "availability_domain",
    "configure_ocpu_name": "shape",
    "oracle_database_software_edition": "database_edition",
    "database_image_oracle_database_version": "db_version",
    "database_name": "db_name",
    "sys_password": "admin_password",
    "ocpu_count": "cpu_core_count",
    "data_storage_gb": "data_storage_size_in_gbs",
    "storage_management_software": "storage_management",
    "storage_performance": "storage_volume_performance_mode",
    "db_system_node_count": "node_count",
    "hostname_prefix": "hostname",
    "license_type": "license_model",
    "ssh_keys": "ssh_public_keys",
    "pdb_name_optional": "pdb_name",
    "db_character_set": "character_set",
    "db_national_character_set": "ncharacter_set",
    "configure_key_management": "key_management",
    "backup_destination": "backup_destination",
    "backup_retention_period": "backup_retention_period",
    "scheduled_day_for_full_backup": "scheduled_day_for_full_backup",
    "scheduled_time_for_daily_backup_utc": "scheduled_time_for_daily_backup_utc",
    "scheduled_time_for_incremental_backup_utc": "scheduled_time_for_incremental_backup_utc",
    "subnet_ocid": "subnet_id",
    "client_subnet_ocid": "subnet_id",
}

LICENSE_MAP = {
    "license included": "LICENSE_INCLUDED",
    "bring your own license": "BRING_YOUR_OWN_LICENSE",
    "byol": "BRING_YOUR_OWN_LICENSE",
}

EDITION_MAP = {
    "enterprise edition high performance": "ENTERPRISE_EDITION_HIGH_PERFORMANCE",
    "enterprise edition extreme performance": "ENTERPRISE_EDITION_EXTREME_PERFORMANCE",
}


def _load_store() -> dict[str, dict[str, Any]]:
    if not os.path.exists(STORE_PATH):
        return {}
    try:
        with open(STORE_PATH, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if isinstance(payload, dict):
            return payload
    except Exception:
        pass
    return {}


def _save_store() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    temp_path = f"{STORE_PATH}.tmp"
    with open(temp_path, "w", encoding="utf-8") as f:
        json.dump(BATCHES, f, indent=2)
    os.replace(temp_path, STORE_PATH)


BATCHES = _load_store()


@app.get("/")
def index():
    with BATCH_LOCK:
        batches = list(BATCHES.values())
        batches.sort(key=lambda b: b["created_at"], reverse=True)
    return render_template("index.html", batches=batches)


@app.post("/upload")
def upload_excel():
    file = request.files.get("excel_file")
    if not file or not file.filename:
        return _redirect_with_message("No Excel file selected.", "error")

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return _redirect_with_message("Please upload a .xlsx or .xlsm file.", "error")

    try:
        rows = _parse_excel(file)
    except Exception as exc:
        return _redirect_with_message(f"Failed to parse Excel: {exc}", "error")

    if not rows:
        return _redirect_with_message("The uploaded sheet has no data rows.", "error")

    batch_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    validation = _validate_rows(rows)

    batch = {
        "id": batch_id,
        "created_at": now,
        "status": "PENDING_REVIEW" if not validation["errors"] else "VALIDATION_ERRORS",
        "approval": "NOT_APPROVED",
        "rows": rows,
        "validation": validation,
        "deploy": {
            "started_at": None,
            "finished_at": None,
            "results": [],
        },
    }

    with BATCH_LOCK:
        BATCHES[batch_id] = batch
        _save_store()

    return redirect(url_for("index"))


@app.post("/approve/<batch_id>")
def approve_batch(batch_id: str):
    with BATCH_LOCK:
        batch = BATCHES.get(batch_id)
        if not batch:
            return _redirect_with_message("Batch not found.", "error")

        if batch["validation"]["errors"]:
            return _redirect_with_message(
                "Cannot approve: batch has validation errors.",
                "error",
            )

        if batch["approval"] == "APPROVED":
            return _redirect_with_message("Batch already approved.", "info")

        batch["approval"] = "APPROVED"
        batch["status"] = "DEPLOYING"
        batch["deploy"]["started_at"] = datetime.utcnow().isoformat()
        _save_store()

    worker = threading.Thread(target=_deploy_batch, args=(batch_id,), daemon=True)
    worker.start()

    return redirect(url_for("index"))


@app.post("/delete/<batch_id>")
def delete_batch(batch_id: str):
    with BATCH_LOCK:
        BATCHES.pop(batch_id, None)
        _save_store()
    return redirect(url_for("index"))


def _parse_excel(file_obj) -> list[dict[str, Any]]:
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    if _looks_like_key_value_template(ws):
        return _parse_key_value_template(ws)

    header_cells = [c.value for c in ws[1]]
    headers = [_normalize_header(h) for h in header_cells]

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue

        item: dict[str, Any] = {}
        for idx, value in enumerate(row):
            if idx >= len(headers):
                continue
            key = headers[idx]
            if not key:
                continue
            item[key] = _clean_cell(value)

        item.setdefault("row_status", "PENDING")
        item.setdefault("row_message", "Awaiting approval")
        rows.append(item)

    return rows


def _looks_like_key_value_template(ws) -> bool:
    sample = 0
    populated_in_col_a = 0
    for r in range(1, min(ws.max_row, 20) + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is None and b is None:
            continue
        sample += 1
        if a is not None and str(a).strip():
            populated_in_col_a += 1
    return sample > 0 and populated_in_col_a >= max(5, int(sample * 0.7))


def _parse_key_value_template(ws) -> list[dict[str, Any]]:
    value_columns = _detect_value_columns(ws)
    if len(value_columns) > 1:
        rows: list[dict[str, Any]] = []
        for col in value_columns:
            col_block = []
            for r in range(1, ws.max_row + 1):
                key = ws.cell(r, 1).value
                value = ws.cell(r, col).value
                col_block.append((key, value))
            mapped = _parse_key_value_block(col_block)
            if _is_meaningful_mapped_row(mapped):
                mapped["_source_column"] = col
                rows.append(mapped)
        return rows

    blocks = _split_key_value_blocks(ws)
    if not blocks:
        return []

    rows = []
    for block in blocks:
        mapped = _parse_key_value_block(block)
        if _is_meaningful_mapped_row(mapped):
            rows.append(mapped)
    return rows


def _detect_value_columns(ws) -> list[int]:
    columns: list[int] = []
    for col in range(2, ws.max_column + 1):
        populated = 0
        anchors = {
            "db_system_name": False,
            "compartment_id": False,
            "db_name": False,
            "shape": False,
            "subnet_id": False,
        }
        for r in range(1, ws.max_row + 1):
            key = ws.cell(r, 1).value
            val = ws.cell(r, col).value
            if key is None or str(key).strip() == "":
                continue
            key_norm = _normalize_header(key)
            if val is not None and str(val).strip() != "":
                populated += 1
                if key_norm == "db_system_name":
                    anchors["db_system_name"] = True
                elif key_norm in {"compartment", "compartment_ocid"}:
                    anchors["compartment_id"] = True
                elif key_norm == "database_name":
                    anchors["db_name"] = True
                elif key_norm == "configure_ocpu_name":
                    anchors["shape"] = True
                elif key_norm in {"subnet_ocid", "subnet_name"}:
                    anchors["subnet_id"] = True
        anchor_count = sum(1 for x in anchors.values() if x)
        if populated >= 5 and anchor_count >= 3:
            columns.append(col)
    return columns


def _split_key_value_blocks(ws) -> list[list[tuple[Any, Any]]]:
    blocks: list[list[tuple[Any, Any]]] = []
    current: list[tuple[Any, Any]] = []

    for r in range(1, ws.max_row + 1):
        key = ws.cell(r, 1).value
        value = ws.cell(r, 2).value

        if key is None or str(key).strip() == "":
            continue

        key_text = str(key).strip().lower()
        is_block_start = key_text.startswith("database basic details")
        if is_block_start and current:
            blocks.append(current)
            current = []

        current.append((key, value))

    if current:
        blocks.append(current)

    return blocks


def _parse_key_value_block(block_rows: list[tuple[Any, Any]]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    section = ""
    for key, value in block_rows:

        if key is None or str(key).strip() == "":
            continue

        key_text = str(key).strip()
        key_norm = _normalize_header(key_text)
        value_is_blank = value is None or str(value).strip() == ""

        # Row-like section labels without values.
        if value_is_blank and (
            key_text.lower().startswith("if ")
            or key_text.lower().endswith("information")
            or key_text.lower().startswith("configure ")
            or key_text.lower().startswith("management")
            or key_text.lower().startswith("encryption")
            or key_text.lower().startswith("provide ")
            or key_text.lower().startswith("database basic details")
        ):
            section = key_norm
            continue

        if value_is_blank:
            continue

        # Keep section-qualified duplicates distinct, and keep latest plain key for simple access.
        if section:
            raw[f"{section}__{key_norm}"] = _clean_cell(value)
        raw[key_norm] = _clean_cell(value)

    mapped: dict[str, Any] = {}
    for key, value in raw.items():
        canonical = KV_TO_CANONICAL_MAP.get(key)
        if canonical:
            mapped[canonical] = value

    _apply_fallback_aliases(raw, mapped)
    _normalize_mapped_values(mapped)
    mapped.setdefault("row_status", "PENDING")
    mapped.setdefault("row_message", "Awaiting approval")
    mapped["_template_type"] = "oracle_key_value"
    return mapped


def _is_meaningful_mapped_row(mapped: dict[str, Any]) -> bool:
    if not mapped:
        return False
    candidates = [
        "display_name",
        "db_name",
        "compartment_id",
        "subnet_id",
        "shape",
        "database_edition",
        "cpu_core_count",
        "data_storage_size_in_gbs",
    ]
    score = sum(1 for c in candidates if str(mapped.get(c, "")).strip())
    return score >= 4


def _apply_fallback_aliases(raw: dict[str, Any], mapped: dict[str, Any]) -> None:
    if "subnet_id" not in mapped:
        for candidate in [
            "subnet_name",
            "client_subnet_in_compartment",
            "if_backup_destination_is_object_storage_fill_below__subnet_name",
        ]:
            if candidate in raw:
                mapped["subnet_id"] = raw[candidate]
                break

    if "pdb_name" not in mapped and "pdb_name_optional" in raw:
        mapped["pdb_name"] = raw["pdb_name_optional"]


def _normalize_mapped_values(mapped: dict[str, Any]) -> None:
    if "license_model" in mapped:
        lm = str(mapped["license_model"]).strip().lower()
        mapped["license_model"] = LICENSE_MAP.get(lm, _normalize_enum_token(mapped["license_model"]))

    if "database_edition" in mapped:
        ed = str(mapped["database_edition"]).strip().lower()
        mapped["database_edition"] = EDITION_MAP.get(ed, _normalize_enum_token(mapped["database_edition"]))

    for key in ["scheduled_time_for_daily_backup_utc", "scheduled_time_for_incremental_backup_utc"]:
        if key in mapped:
            mapped[key] = _excel_time_to_hhmm(mapped[key])

    if str(mapped.get("pdb_name", "")).strip().upper() == str(mapped.get("db_name", "")).strip().upper():
        mapped["pdb_name"] = f"{str(mapped['db_name']).strip()}PDB"

    if "storage_management" in mapped:
        sm = str(mapped["storage_management"]).strip().lower()
        if "logical" in sm or sm == "lvm":
            mapped["storage_management"] = "LVM"
        elif "grid" in sm or sm == "asm":
            mapped["storage_management"] = "ASM"
        else:
            mapped["storage_management"] = _normalize_enum_token(mapped["storage_management"])

    if "storage_volume_performance_mode" in mapped:
        sp = str(mapped["storage_volume_performance_mode"]).strip().lower()
        if sp == "balanced":
            mapped["storage_volume_performance_mode"] = "BALANCED"
        elif "high" in sp:
            mapped["storage_volume_performance_mode"] = "HIGH_PERFORMANCE"
        else:
            mapped["storage_volume_performance_mode"] = _normalize_enum_token(mapped["storage_volume_performance_mode"])


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
        .replace("(", "")
        .replace(")", "")
    )


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _excel_time_to_hhmm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        total_minutes = round(float(value) * 24 * 60)
        hours = (total_minutes // 60) % 24
        mins = total_minutes % 60
        return f"{hours:02d}:{mins:02d}"
    text = str(value).strip()
    if not text:
        return ""
    # Handle "2:00 PM" style values if present.
    try:
        parts = text.split(":")
        if len(parts) >= 2:
            h = int(parts[0]) % 24
            m = int(parts[1][:2])
            return f"{h:02d}:{m:02d}"
    except Exception:
        pass
    return text


def _normalize_enum_token(value: Any) -> str:
    text = str(value).strip()
    if not text:
        return ""
    text = re.sub(r"[^A-Za-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text.upper()


def _validate_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    errors = []
    warnings = []

    for i, row in enumerate(rows, start=2):
        for col in REQUIRED_COLUMNS:
            if str(row.get(col, "")).strip() == "":
                errors.append(f"Row {i}: Missing required column '{col}'")

        cpu = row.get("cpu_core_count")
        storage = row.get("data_storage_size_in_gbs")
        try:
            if cpu != "":
                int(cpu)
        except Exception:
            errors.append(f"Row {i}: cpu_core_count must be a number")

        try:
            if storage != "":
                int(storage)
        except Exception:
            errors.append(f"Row {i}: data_storage_size_in_gbs must be a number")

        if row.get("admin_password") and len(str(row["admin_password"])) < 9:
            warnings.append(f"Row {i}: admin_password seems too short")

        if row.get("compartment_id") and not OCID_PATTERN.match(str(row["compartment_id"]).strip()):
            errors.append(f"Row {i}: compartment_id must be an OCID (ocid1...)")

        if row.get("subnet_id") and not OCID_PATTERN.match(str(row["subnet_id"]).strip()):
            errors.append(f"Row {i}: subnet_id must be an OCID (ocid1...)")

        if not str(row.get("availability_domain", "")).strip():
            warnings.append(f"Row {i}: availability_domain not provided; will auto-select during deployment")

        for k in ["scheduled_time_for_daily_backup_utc", "scheduled_time_for_incremental_backup_utc"]:
            t = str(row.get(k, "")).strip()
            if t and not re.match(r"^([01]\d|2[0-3]):[0-5]\d$", t):
                warnings.append(f"Row {i}: {k} should be HH:MM (UTC), got '{t}'")

    return {"errors": errors, "warnings": warnings}


def _deploy_batch(batch_id: str) -> None:
    try:
        with BATCH_LOCK:
            batch = BATCHES.get(batch_id)
            if not batch:
                return
            rows = list(batch["rows"])

        results = []
        final_status = "DEPLOYED"

        for idx, row in enumerate(rows, start=1):
            try:
                deploy_result = _deploy_single_dbcs(row)
                results.append({"row": idx, **deploy_result})
            except Exception as exc:
                final_status = "DEPLOYED_WITH_ERRORS"
                results.append(
                    {
                        "row": idx,
                        "status": "FAILED",
                        "message": str(exc),
                        "db_system_id": "",
                    }
                )

        with BATCH_LOCK:
            batch = BATCHES.get(batch_id)
            if not batch:
                return
            batch["deploy"]["results"] = results
            batch["deploy"]["finished_at"] = datetime.utcnow().isoformat()
            batch["status"] = final_status
            _save_store()
    except Exception as exc:
        with BATCH_LOCK:
            batch = BATCHES.get(batch_id)
            if not batch:
                return
            batch["deploy"]["results"] = [
                {
                    "row": 0,
                    "status": "FAILED",
                    "message": f"Deployment worker crashed: {exc}",
                    "db_system_id": "",
                }
            ]
            batch["deploy"]["finished_at"] = datetime.utcnow().isoformat()
            batch["status"] = "DEPLOYED_WITH_ERRORS"
            _save_store()


def _deploy_single_dbcs(row: dict[str, Any]) -> dict[str, str]:
    generated_keys = _generate_ssh_key_pair(row.get("display_name", "dbcs"))
    ssh_public_keys = [generated_keys["public_key"]]

    dry_run = os.getenv("DRY_RUN", "true").lower() == "true"
    if dry_run:
        fake_id = f"ocid1.dbsystem.oc1..simulated-{uuid.uuid4().hex[:12]}"
        return {
            "status": "SUCCESS",
            "message": (
                "DRY_RUN enabled; deployment simulated. "
                f"SSH keys generated at {generated_keys['private_key_path']} and {generated_keys['public_key_path']}."
            ),
            "db_system_id": fake_id,
        }

    try:
        import oci
    except ImportError as exc:
        raise RuntimeError("OCI SDK is not installed. Install requirements first.") from exc

    profile = os.getenv("OCI_PROFILE", "DEFAULT")
    config = _load_oci_config(profile)
    connect_timeout = float(os.getenv("OCI_CONNECT_TIMEOUT", "10"))
    read_timeout = float(os.getenv("OCI_READ_TIMEOUT", "45"))
    client = oci.database.DatabaseClient(config, timeout=(connect_timeout, read_timeout))
    availability_domain = _resolve_availability_domain(row, config)

    payload = {
        "compartmentId": row["compartment_id"],
        "availabilityDomain": availability_domain,
        "subnetId": row["subnet_id"],
        "shape": row["shape"],
        "cpuCoreCount": int(row["cpu_core_count"]),
        "databaseEdition": row["database_edition"],
        "displayName": row["display_name"],
        "hostname": row.get("hostname") or _default_hostname(row["display_name"]),
        "nodeCount": int(row.get("node_count", 1) or 1),
        "dataStorageSizeInGBs": int(row["data_storage_size_in_gbs"]),
        "licenseModel": row.get("license_model", "LICENSE_INCLUDED") or "LICENSE_INCLUDED",
        "sshPublicKeys": ssh_public_keys,
        "nsgIds": _split_csv(row.get("nsg_ids", "")),
        "dbHome": {
            "dbVersion": row["db_version"],
            "database": {
                "dbName": row["db_name"],
                "adminPassword": row["admin_password"],
                "pdbName": row.get("pdb_name") or "PDB1",
                "dbWorkload": row.get("db_workload", "OLTP") or "OLTP",
                "characterSet": row.get("character_set", "AL32UTF8") or "AL32UTF8",
                "ncharacterSet": row.get("ncharacter_set", "AL16UTF16") or "AL16UTF16",
                "dbBackupConfig": {
                    "autoBackupEnabled": _to_bool(row.get("auto_backup_enabled", "false")),
                },
            },
        },
    }

    storage_mgmt = str(row.get("storage_management", "")).strip()
    if storage_mgmt:
        payload["dbSystemOptions"] = {"storageManagement": storage_mgmt}

    storage_perf = str(row.get("storage_volume_performance_mode", "")).strip()
    if storage_perf:
        payload["storageVolumePerformanceMode"] = storage_perf

    response = client.launch_db_system(payload)
    db_system_id = getattr(response.data, "id", "")

    return {
        "status": "SUCCESS",
        "message": (
            f"Provisioning request submitted to OCI (AD: {availability_domain}). "
            f"Requested storage management: {storage_mgmt or 'OCI default'}. "
            f"SSH keys saved at {generated_keys['private_key_path']} and {generated_keys['public_key_path']}."
        ),
        "db_system_id": db_system_id,
    }


def _resolve_availability_domain(row: dict[str, Any], config: dict[str, Any]) -> str:
    ad = str(row.get("availability_domain", "")).strip()
    if ad:
        return ad

    import oci

    connect_timeout = float(os.getenv("OCI_CONNECT_TIMEOUT", "10"))
    read_timeout = float(os.getenv("OCI_READ_TIMEOUT", "45"))
    identity = oci.identity.IdentityClient(config, timeout=(connect_timeout, read_timeout))
    tenancy_id = config.get("tenancy")
    if not tenancy_id:
        raise RuntimeError("OCI config missing tenancy OCID; cannot auto-select availability domain")

    response = identity.list_availability_domains(compartment_id=tenancy_id)
    ads = sorted(response.data or [], key=lambda x: getattr(x, "name", ""))
    if not ads:
        raise RuntimeError("No availability domains found in tenancy for this region")

    return ads[0].name


def _load_oci_config(profile: str) -> dict[str, Any]:
    try:
        import oci
    except ImportError as exc:
        raise RuntimeError("OCI SDK is not installed. Install requirements first.") from exc

    config_path = os.getenv("OCI_CLI_CONFIG_FILE", os.path.expanduser("~/.oci/config"))
    parser = configparser.ConfigParser()

    if not os.path.exists(config_path):
        raise RuntimeError(f"OCI config file not found: '{config_path}'")

    parser.read(config_path)
    if profile not in parser:
        raise RuntimeError(f"OCI profile '{profile}' not found in '{config_path}'")

    config = dict(parser[profile])
    key_file = str(config.get("key_file", "")).strip()
    if key_file:
        key_file = os.path.expandvars(os.path.expanduser(key_file))
        if not os.path.exists(key_file):
            fixed = _maybe_fix_windows_key_path(key_file)
            if fixed:
                key_file = fixed
        config["key_file"] = key_file

    final_key = str(config.get("key_file", "")).strip()
    if final_key and not os.path.exists(final_key):
        raise RuntimeError(
            f"OCI key file not found: '{final_key}'. If running in Docker, map ~/.oci to /root/.oci "
            "and set key_file accordingly."
        )

    try:
        oci.config.validate_config(config)
    except Exception as exc:
        raise RuntimeError(f"OCI config validation failed for profile '{profile}': {exc}") from exc

    return config


def _maybe_fix_windows_key_path(path_value: str) -> str:
    # If config contains a Windows path like C:\\Users\\...\\.oci\\oci_api_key.pem
    # but app runs in Linux container, try common mounted .oci paths.
    if "\\" not in path_value and ":" not in path_value:
        return ""

    filename = ntpath.basename(path_value)
    candidates = [
        os.path.join("/root/.oci", filename),
        os.path.join(os.path.expanduser("~/.oci"), filename),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return ""


def _split_csv(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"[\r\n,]+", text)
    return [x.strip() for x in parts if x.strip()]


def _generate_ssh_key_pair(dbcs_name: str) -> dict[str, str]:
    try:
        from cryptography.hazmat.primitives import serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
    except ImportError as exc:
        raise RuntimeError(
            "cryptography package is required for SSH key generation. Install requirements.txt first."
        ) from exc

    safe_name = _safe_filename(dbcs_name or "dbcs")
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    key_dir = os.path.join(os.getcwd(), "keys", safe_name)
    os.makedirs(key_dir, exist_ok=True)

    private_key_path = os.path.join(key_dir, f"{safe_name}_{stamp}")
    public_key_path = f"{private_key_path}.pub"

    private_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    private_bytes = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.TraditionalOpenSSL,
        encryption_algorithm=serialization.NoEncryption(),
    )
    public_bytes = private_key.public_key().public_bytes(
        encoding=serialization.Encoding.OpenSSH,
        format=serialization.PublicFormat.OpenSSH,
    )

    with open(private_key_path, "wb") as f:
        f.write(private_bytes)
    with open(public_key_path, "wb") as f:
        f.write(public_bytes + b"\n")

    # Best-effort on non-POSIX systems.
    try:
        os.chmod(private_key_path, 0o600)
    except Exception:
        pass

    return {
        "private_key_path": private_key_path,
        "public_key_path": public_key_path,
        "public_key": public_bytes.decode("utf-8").strip(),
    }


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "-", value.strip())
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-")
    return cleaned[:64] or "dbcs"


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _default_hostname(display_name: str) -> str:
    safe = "".join(ch.lower() if ch.isalnum() else "-" for ch in display_name)
    safe = "-".join(filter(None, safe.split("-")))
    return (safe[:15] or "dbcs-host")


def _redirect_with_message(message: str, level: str):
    return redirect(url_for("index", msg=message, level=level))


@app.context_processor
def inject_request_context():
    message = request.args.get("msg", "")
    level = request.args.get("level", "info")
    return {"flash_message": message, "flash_level": level}


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8080"))
    app.run(host="0.0.0.0", port=port, debug=True)
